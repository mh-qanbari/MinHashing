from openpyxl import load_workbook


class ExcelManager:
    def __init__(self, filename):
        """ Initializes instance of the class.
        :param filename: Excel file path.
        :type filename: str
        """
        __workbook = load_workbook(filename, read_only=True, data_only=True)
        self.__sheet = __workbook.get_sheet_by_name(__workbook.get_sheet_names()[0])
        del __workbook

    def __getitem__(self, t):
        """ Returns content of the cell which is in the 't' position.
        :param t: Cell position e.g. t=(row, col) (row and col are 1-based indexes)
        :type t: tuple(2)
        :return: The specified cell content.
        :rtype: any
        """
        if type(t) is not tuple or len(t) != 2:
            print "[EXCEPTION] @ExcelManager.__getitem__(t) > Wrong input format. 't' must be a tuple of size 2."
            return None
        return self.__sheet.cell(row=t[0], column=t[1]).value

    def getHeader(self, col=1):
        return self[1, col]

    def rows(self):
        return self.__sheet.iter_rows(row_offset=1)
        # __ignore_first = True
        # for __row in self.__sheet.rows:
        #     if __ignore_first:
        #         continue
        #     yield [(__row[i]) for i in range()]

    # [HW3] : Special for 3`rd series homework
    def getSellers(self):
        for __row in self.__sheet.iter_rows(row_offset=1):
            yield __row[2].value

    def getRegionsGoods(self):
        for __row in self.__sheet.iter_rows(row_offset=1):
            yield __row[3].value, __row[1].value

    def getRegionsGoodsbyMonth(self):
        for __row in self.__sheet.iter_rows(row_offset=1):
            yield __row[0].value.strftime("%b"), __row[1].value, __row[3].value
    # [HW3]; ----------------------------------
